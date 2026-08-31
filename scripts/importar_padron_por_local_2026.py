#!/usr/bin/env python3
"""Importa el padrón municipal 2026 desde XLSX organizados por local.

Uso de prueba (no modifica la base):
    python scripts/importar_padron_por_local_2026.py "C:\\ruta\\DEPARTAMENTO 11-CENTRAL"

Ejecución real:
    python scripts/importar_padron_por_local_2026.py "C:\\ruta\\DEPARTAMENTO 11-CENTRAL" --execute
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterator

try:
    import openpyxl
except ImportError:
    raise SystemExit("Falta openpyxl. Instale con: py -m pip install openpyxl mysql-connector-python")


HEADING_RE = re.compile(
    r"Departamento:\s*\d+-(.*?)\s*\|\s*"
    r"Distrito:\s*\d+-(.*?)\s*\|\s*"
    r"Local:\s*\d+-(.*)$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class LocalFile:
    path: Path
    departamento: str
    distrito: str
    local: str
    filas: int
    mesas: int
    digest: str

    @property
    def key(self) -> tuple[str, str, str]:
        return (self.departamento, self.distrito, self.local)


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def parse_heading(value) -> tuple[str, str, str]:
    match = HEADING_RE.search(clean(value))
    if not match:
        raise ValueError(f"Encabezado no reconocido: {value!r}")
    return tuple(clean(match.group(i)) for i in range(1, 4))


def split_name(value) -> tuple[str, str]:
    full_name = clean(value)
    if "," not in full_name:
        return "", full_name
    apellido, nombre = full_name.split(",", 1)
    return clean(apellido), clean(nombre)


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    text = clean(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Fecha no reconocida: {text!r}")


def iter_voters(path: Path) -> Iterator[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        departamento, distrito, local = parse_heading(sheet.cell(2, 1).value)
        for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
            mesa, orden, cedula = row[0], row[1], row[2]
            if cedula in (None, ""):
                continue
            try:
                mesa = int(mesa)
                orden = int(orden)
                cedula = str(int(cedula)) if isinstance(cedula, float) else clean(cedula)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{path.name}, fila {row_number}: cédula/mesa/orden inválidos") from exc
            apellido, nombre = split_name(row[3])
            yield (
                cedula,
                nombre,
                apellido,
                parse_date(row[4]),
                clean(row[5]) or None,
                clean(row[8]) or None,
                departamento,
                distrito,
                local,
                local,
                mesa,
                orden,
                path.name,
            )
    finally:
        workbook.close()


def inspect_file(path: Path) -> LocalFile:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        departamento, distrito, local = parse_heading(sheet.cell(2, 1).value)
        filas = 0
        mesas = set()
        digest = hashlib.sha256()
        for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
            if row[2] in (None, ""):
                continue
            try:
                mesa, orden = int(row[0]), int(row[1])
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{path.name}, fila {row_number}: mesa/orden inválidos") from exc
            cedula = str(int(row[2])) if isinstance(row[2], float) else clean(row[2])
            digest.update(f"{mesa}|{orden}|{cedula}\n".encode("utf-8"))
            mesas.add(mesa)
            filas += 1
        if not filas:
            raise ValueError(f"Archivo sin electores: {path}")
        return LocalFile(path, departamento, distrito, local, filas, len(mesas), digest.hexdigest())
    finally:
        workbook.close()


def read_env(path: Path) -> dict[str, str]:
    values = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def db_config(project: Path) -> dict:
    env = read_env(project / ".env")
    return {
        "host": env.get("DB_HOST", "127.0.0.1"),
        "port": int(env.get("DB_PORT", "3306")),
        "user": env.get("DB_USERNAME", "root"),
        "password": env.get("DB_PASSWORD", ""),
        "database": env.get("DB_DATABASE", "eleccionnew"),
        "charset": "utf8mb4",
        "use_unicode": True,
    }


def batches(iterator: Iterator[tuple], size: int = 2000) -> Iterator[list[tuple]]:
    batch = []
    for item in iterator:
        batch.append(item)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def validate_source(source: Path) -> tuple[list[LocalFile], list[LocalFile]]:
    paths = sorted(source.rglob("*.xlsx"))
    if not paths:
        raise ValueError(f"No se encontraron XLSX en {source}")

    unique: dict[tuple[str, str, str], LocalFile] = {}
    duplicates = []
    for index, path in enumerate(paths, 1):
        info = inspect_file(path)
        previous = unique.get(info.key)
        if previous:
            if previous.digest != info.digest or previous.filas != info.filas:
                raise ValueError(
                    "Dos archivos distintos declaran el mismo local:\n"
                    f"  {previous.path}\n  {info.path}"
                )
            duplicates.append(info)
        else:
            unique[info.key] = info
        print(f"\rValidando XLSX: {index}/{len(paths)}", end="", flush=True)
    print()
    return sorted(unique.values(), key=lambda x: x.key), duplicates


def print_summary(files: list[LocalFile], duplicates: list[LocalFile]) -> None:
    by_district = defaultdict(lambda: [0, 0, 0])
    for item in files:
        data = by_district[(item.departamento, item.distrito)]
        data[0] += 1
        data[1] += item.mesas
        data[2] += item.filas

    print("\nResumen validado")
    print("-" * 76)
    for (departamento, distrito), (locales, mesas, filas) in sorted(by_district.items()):
        print(f"{departamento:12} | {distrito:25} | {locales:3} locales | {mesas:4} mesas | {filas:8} electores")
    print("-" * 76)
    print(f"Total: {len(files)} locales, {sum(x.mesas for x in files)} mesas, {sum(x.filas for x in files)} electores")
    print(f"Duplicados idénticos omitidos: {len(duplicates)}")


def import_database(files: list[LocalFile], project: Path) -> None:
    try:
        import mysql.connector
    except ImportError:
        raise SystemExit("Falta mysql-connector-python. Instale con: py -m pip install mysql-connector-python")

    config = db_config(project)
    connection = mysql.connector.connect(**config)
    cursor = connection.cursor()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_padron = f"padron_nuevo_{stamp}"
    new_locales = f"locales_internas_nuevo_{stamp}"
    backup_padron = f"padron_backup_{stamp}"
    backup_locales = f"locales_internas_backup_{stamp}"

    insert_padron = f"""
        INSERT INTO `{new_padron}`
        (cedula,nombre,apellido,fecha_nacimiento,afiliaciones,tipo,
         departamento_nombre,distrito_nombre,local_generales,local_interna,
         mesa,orden,archivo_origen)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """

    try:
        cursor.execute(f"CREATE TABLE `{new_padron}` LIKE `padron`")
        cursor.execute(f"CREATE TABLE `{new_locales}` LIKE `locales_internas`")
        connection.commit()

        expected = sum(item.filas for item in files)
        imported = 0
        local_rows = []
        seen_cedulas = Counter()

        for file_number, item in enumerate(files, 1):
            for batch in batches(iter_voters(item.path)):
                cursor.executemany(insert_padron, batch)
                imported += len(batch)
                seen_cedulas.update(row[0] for row in batch)
            local_rows.append((item.distrito, item.departamento, item.local, item.mesas))
            connection.commit()
            print(
                f"\rImportando locales: {file_number}/{len(files)} "
                f"({imported:,}/{expected:,} electores)",
                end="",
                flush=True,
            )
        print()

        cursor.executemany(
            f"INSERT INTO `{new_locales}` "
            "(distrito_nombre,departamento_nombre,local_interna,cantmesa) VALUES (%s,%s,%s,%s)",
            local_rows,
        )
        connection.commit()

        cursor.execute(f"SELECT COUNT(*),COUNT(DISTINCT cedula) FROM `{new_padron}`")
        count_rows, count_cedulas = cursor.fetchone()
        cursor.execute(f"SELECT COUNT(*) FROM `{new_locales}`")
        count_locales = cursor.fetchone()[0]
        if count_rows != expected or count_locales != len(files):
            raise RuntimeError("Los conteos importados no coinciden con los XLSX; no se realizó el reemplazo")

        duplicate_cedulas = sum(1 for count in seen_cedulas.values() if count > 1)
        print(f"Validación: {count_rows:,} filas; {count_cedulas:,} cédulas únicas; {count_locales} locales")
        print(f"Cédulas repetidas en los XLSX: {duplicate_cedulas}")

        cursor.execute(
            f"RENAME TABLE "
            f"`padron` TO `{backup_padron}`, `{new_padron}` TO `padron`, "
            f"`locales_internas` TO `{backup_locales}`, `{new_locales}` TO `locales_internas`"
        )
        connection.commit()
        print("\nReemplazo completado correctamente.")
        print(f"Respaldo del padrón anterior: {backup_padron}")
        print(f"Respaldo de locales anteriores: {backup_locales}")
    except Exception:
        connection.rollback()
        for table in (new_padron, new_locales):
            try:
                cursor.execute(f"DROP TABLE IF EXISTS `{table}`")
            except Exception:
                pass
        connection.commit()
        raise
    finally:
        cursor.close()
        connection.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Importar padrón 2026 por local a Laravel/MySQL")
    parser.add_argument("source", type=Path, help="Carpeta DEPARTAMENTO 11-CENTRAL")
    parser.add_argument("--execute", action="store_true", help="Realizar el reemplazo en MySQL")
    parser.add_argument("--expected-locales", type=int, default=241)
    parser.add_argument("--expected-electores", type=int, default=1_384_092)
    parser.add_argument(
        "--allow-incomplete",
        action="store_true",
        help="Permitir reemplazo aunque falten locales/electores (no recomendado)",
    )
    parser.add_argument(
        "--project",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Proyecto Laravel que contiene .env",
    )
    args = parser.parse_args()

    source = args.source.resolve()
    if not source.is_dir():
        print(f"No existe la carpeta: {source}", file=sys.stderr)
        return 2

    try:
        files, duplicates = validate_source(source)
        print_summary(files, duplicates)
        actual_rows = sum(item.filas for item in files)
        complete = len(files) == args.expected_locales and actual_rows == args.expected_electores
        if not complete:
            print(
                "\nADVERTENCIA: el conjunto está incompleto: "
                f"se esperaban {args.expected_locales} locales y {args.expected_electores:,} electores."
            )
            if args.execute and not args.allow_incomplete:
                raise RuntimeError(
                    "Se canceló el reemplazo para proteger la base. "
                    "Copie los XLSX faltantes o use --allow-incomplete bajo su responsabilidad."
                )
        if not args.execute:
            print("\nModo validación: la base de datos NO fue modificada.")
            print("Si el resumen es correcto, repita el comando agregando --execute")
            return 0
        import_database(files, args.project.resolve())
        return 0
    except Exception as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
